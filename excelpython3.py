from openpyxl import Workbook #load_workbook
from openpyxl.utils import get_column_letter
wb = Workbook()
ws = wb.active

for row in range(1,11):         #allows you to loop through the rows in aa sheet
    for col in range(1,7):      # allos you to loop thorugh the coloumns in a sheet
        #char = chr(65 + col) # 65 = character A
        char = get_column_letter(col)
        ws[char + str(row)] = char +str(row)

wb.save('character_values.xlsx')