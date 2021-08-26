from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
wb = load_workbook('tim.xlsx')
ws = wb.active

for row in range(1,11):
    for col in range(1,5):
        #char = chr(65 + col) # 65 = character A
        char = get_column_letter(col)
        print(ws[char + str(row)].value)